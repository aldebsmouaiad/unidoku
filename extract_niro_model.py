#!/usr/bin/env python3
"""
Extract NIRO Reifegradmodell (Technische Dokumentation) from Excel into JSON.

- Reads an Excel workbook with one sheet per Dimension (e.g., TD1.1, OG4.2, ...)
- Extracts:
  * Dimension meta (code, name, category)
  * Prozess-Steckbrief (Zweck, Ergebnisse, Basispraktiken, Arbeitsprodukte)
  * Fragen je Reifegradstufe (1..5)
  * Abnahmekriterien je Stufe
  * Nutzen bei Erreichen der Stufe je Stufe
  * Glossar (Term -> Definition) from sheet "Glossar"

No runtime data is stored by the Streamlit app; this script only generates the model config JSON.

Usage:
  python extract_niro_model.py \
      --input "/path/to/20251209_NIRO_Reifegradmodell_Technische_Dokumentation.xlsx" \
      --output "/path/to/unidoku/data/models/niro_td_model.json"
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import re
from typing import Any, Dict, List

import openpyxl


LEVELS_INFO = {
    1: "initial",
    2: "gemanagt",
    3: "definiert",
    4: "quantitativ gemanagt",
    5: "optimiert",
}


def clean_text(v: Any) -> str:
    """Normalize cell values into stable strings while preserving newlines."""
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    s = str(v).replace("\u00a0", " ").strip()
    s = re.sub(r"[ \t]+", " ", s)
    s = "\n".join(line.strip() for line in s.splitlines())
    return s.strip()


def find_row_with_value(ws: openpyxl.worksheet.worksheet.Worksheet, value: str, col: str | None = None, max_row: int = 200) -> int | None:
    """Find first row where a cell equals `value`. Optionally restrict to a column letter."""
    max_r = min(max_row, ws.max_row)
    if col:
        for r in range(1, max_r + 1):
            v = ws[f"{col}{r}"].value
            if isinstance(v, str) and v.strip() == value:
                return r
        return None

    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == value:
                return r
    return None


def extract_process_profile(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, str]:
    """
    Extract Prozess-Steckbrief fields from a Dimension sheet.

    Expected layout (like in the Excel):
      Col B: labels  (Zweck, Ergebnisse, Basispraktiken, Arbeitsprodukte)
      Col C: texts
    """
    fields: Dict[str, str] = {}
    mapping = [
        ("Zweck", "purpose"),
        ("Ergebnisse", "results"),
        ("Basispraktiken", "basic_practices"),
        ("Arbeitsprodukte", "work_products"),
    ]
    for label, key in mapping:
        r = find_row_with_value(ws, label, col="B", max_row=30)
        fields[key] = clean_text(ws[f"C{r}"].value) if r else ""
    return fields


def extract_dimension(ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> Dict[str, Any]:
    # Header: code/name in row 6, cols B/C (as seen in the workbook)
    code = clean_text(ws["B6"].value) or sheet_name
    name = clean_text(ws["C6"].value)
    m = re.match(r"^[A-Z]{2}", code)
    category = m.group(0) if m else ""

    process_profile = extract_process_profile(ws)
    description = process_profile.get("purpose", "")

    # Questions table header starts at a row where col B contains "Zu Reifegrad"
    header_row = find_row_with_value(ws, "Zu Reifegrad", col="B", max_row=80)
    if header_row is None:
        raise ValueError(f"Header 'Zu Reifegrad' not found in sheet {sheet_name!r}")

    data_start = header_row + 2  # in this workbook: header at row 14, data starts at row 16

    # Last row determined by last non-empty question cell in column C
    last_row = 0
    for r in range(data_start, ws.max_row + 1):
        if ws.cell(r, 3).value not in (None, ""):
            last_row = r
    if last_row == 0:
        last_row = data_start - 1

    levels: Dict[int, Dict[str, Any]] = {}
    current_level: int | None = None
    q_counts: Dict[int, int] = {}

    for r in range(data_start, last_row + 1):
        lv = ws[f"B{r}"].value
        if lv is not None and str(lv).strip() != "":
            try:
                lv_int = int(float(lv))
                if 1 <= lv_int <= 5:
                    current_level = lv_int
            except ValueError:
                pass

        q_text = clean_text(ws[f"C{r}"].value)
        if not q_text or current_level is None:
            continue

        level_obj = levels.setdefault(
            current_level,
            {
                "level_number": current_level,
                "name": LEVELS_INFO.get(current_level, str(current_level)),
                "acceptance_criteria": "",
                "benefit": "",
                "questions": [],
            },
        )

        # These are usually merged cells per level; only the top-left contains the value in openpyxl.
        acc = clean_text(ws[f"F{r}"].value)  # Abnahmekriterien
        ben = clean_text(ws[f"J{r}"].value)  # Nutzen bei Erreichen der Stufe
        if acc and not level_obj["acceptance_criteria"]:
            level_obj["acceptance_criteria"] = acc
        if ben and not level_obj["benefit"]:
            level_obj["benefit"] = ben

        q_counts[current_level] = q_counts.get(current_level, 0) + 1
        q_id = f"{code}-L{current_level}-Q{q_counts[current_level]}"
        level_obj["questions"].append({"id": q_id, "text": q_text})

    level_list: List[Dict[str, Any]] = [levels[k] for k in sorted(levels.keys())]

    return {
        "code": code,
        "name": name,
        "category": category,
        "description": description,
        "default_target_level": 3,
        "process_profile": process_profile,
        "levels": level_list,
    }


def extract_glossary(wb: openpyxl.Workbook) -> Dict[str, str]:
    """Extract glossary entries from sheet 'Glossar' (term in col C, definition in col E)."""
    if "Glossar" not in wb.sheetnames:
        return {}

    ws = wb["Glossar"]
    glossary: Dict[str, str] = {}

    for r in range(1, ws.max_row + 1):
        term = clean_text(ws[f"C{r}"].value)
        definition = clean_text(ws[f"E{r}"].value)

        if not term or not definition:
            continue

        # Skip single-letter headings ("A", "B", ...)
        if len(term) == 1 and term.isalpha():
            continue

        glossary[term] = definition

    return glossary


def extract_model(input_path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)

    title = ""
    if "Deckblatt" in wb.sheetnames:
        title = clean_text(wb["Deckblatt"]["C9"].value)

    description_parts: List[str] = []
    if "Einführung" in wb.sheetnames:
        ws = wb["Einführung"]
        for r in range(6, 25):
            v = ws[f"B{r}"].value
            if v:
                description_parts.append(clean_text(v))

    model_description = " ".join(description_parts).strip()

    glossary = extract_glossary(wb)

    dim_sheets = [name for name in wb.sheetnames if re.match(r"^[A-Z]{2}\d+\.\d+$", name)]
    dimensions: List[Dict[str, Any]] = []
    for sh in dim_sheets:
        dimensions.append(extract_dimension(wb[sh], sh))

    return {
        "name": title or "NIRO Reifegradmodell Technische Dokumentation",
        "description": model_description,
        "levels_info": {str(k): v for k, v in LEVELS_INFO.items()},
        "glossary": glossary,
        "dimensions": dimensions,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to the NIRO Excel workbook (.xlsx)")
    parser.add_argument("--output", required=True, help="Path to write the model JSON")
    args = parser.parse_args()

    model = extract_model(args.input)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(model, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(model['dimensions'])} dimensions, {len(model['glossary'])} glossary terms -> {args.output}")


if __name__ == "__main__":
    main()
